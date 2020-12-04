# https://xlsxwriter.readthedocs.io/
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import re


class To_excel():

    def __init__(self, workbook_file_name, sheet_name):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.get_sheet_by_name("Sheet") )
        self.workbook.create_sheet(title = sheet_name)
        self.file_name = workbook_file_name
        self.workbook.save(filename = workbook_file_name)
    pass

    def __write_headers(self, n_features):
        sheet = self.workbook.get_sheet_by_name("84_vanila")
        model_list = ["DT", "GP", "KNN", "MLP", "NB", "RF", "svm"]
        for row in [2, 10, 18]:
            sheet.cell(row = row, column = 1).value = "Dataset"
            sheet.cell(row = row, column = 2).value = "Instância"
            column_step = (n_features - 1)
            start_merge = 3
            for model in model_list:
                sheet.cell(row = row, column = start_merge).value = model
                sheet.cell(row = row, column = start_merge).alignment = Alignment(horizontal = "center",
                                                                                  vertical = "center")
                sheet.merge_cells(start_row = row,
                                    start_column = start_merge,
                                    end_row = row,
                                    end_column = 3 + column_step)
                start_merge = start_merge + n_features
                column_step = column_step + (n_features - 1) + 1

        self.workbook.save(self.file_name)
    pass

    def __write_colum_dataset(self, experiment_list, n_features):

        columns = ['Dataset', 'Instance', 'Conf.', 'Model', 'feature_name', 'feature_weigth']
        data = []
        for feat in experiment_list:
            predicted_prob, class_name, feature_importance = feat.get_prob_and_feat_list()
            # print("------------------------------------------------")
            # print(feat.get_dataset())
            # print(feat.get_instance_index())
            # print(feat.get_class_inst_baseline())
            # print(feat.get_model())

            for fi in feature_importance:
                data.append([feat.get_dataset(), feat.get_instance_index(),
                             feat.get_class_inst_baseline(), feat.get_model(),
                             fi.feature_name, fi.feature_weigth
                             ])

        df_ = pd.DataFrame(data = data, columns = columns)
        df_.sort_values(by=['Model', 'Dataset'], inplace=True)

        # print(df_.feature_name)
        # ma = re.search("<=", df_.feature_name.str)

        # df_['menor_igual'] = df_.feature_name.str.match(".*<(?=)")
        # df_['maior_igual'] = df_.feature_name.str.match(".*>(?=)")
        # df_['menor'] = df_.feature_name.str.match(".*<(?!=)")
        # df_['maior'] = df_.feature_name.str.match(".*>(?!=)")

        df_['menor_igual'] = df_.feature_name.str.match(".*<={1}")
        df_['maior_igual'] = df_.feature_name.str.match(".*>={1}")
        df_['menor'] = df_.feature_name.str.match(".*< {1}")
        df_['maior'] = df_.feature_name.str.match(".*> {1}")
        df_['between'] = df_.feature_name.str.match(".*<.*<=")

        # for column in df_[['feature_name']]:
        #     print('Column Contents : ', column)
        # df_ = df_.loc[df_.Dataset == "heart",:]
        # for index, row in df_[['feature_name']].iterrows():
        #     print(re.sub("[^a-zA-Z]", "", row.feature_name))
        #
        #     if len(re.findall("""<""", row.feature_name)) == 2:
        #         print(re.findall("""<""", row.feature_name))
        #
        #     if len(re.findall("""<""", row.feature_name)) == 1:
        #         print(re.findall("""<""", row.feature_name))
        #
        #     if len(re.findall("""<""", row.feature_name)) == 0:
        #         print(re.findall("""<""", row.feature_name))

            # print(re.findall(""">""", row.feature_name))

        # print(df_)
        # df_.to_excel("output.xlsx"xlsx)
        #
        #
        # tmp_diabetes = df_.loc[df_.Dataset == "diabetes2", :]
        # tmp_diabetes = tmp_diabetes.pivot(index=['Dataset', 'Model', 'Conf.'],
        #                                     columns=['feature_name'],
        #                                     values=['feature_weigth'])
        #
        # tmp_diabetes.to_excel("tmp_diabetes.xlsx")
        #
        # #x
        # tmp_heart = df_.loc[df_.Dataset == "heart", :]
        # tmp_heart = tmp_heart.pivot(index=['Dataset', 'Model', 'Conf.'],
        #                                     columns=['feature_name'],
        #                                     values=['feature_weigth'])
        #
        # tmp_heart.to_excel("tmp_heart.xlsx")
        #
        # #x
        # tmp_wine2 = df_.loc[df_.Dataset == "wine2", :]
        # tmp_wine2 = tmp_wine2.pivot(index=['Dataset', 'Model', 'Conf.'],
        #                                     columns=['feature_name'],
        #                                     values=['feature_weigth'])
        #
        # tmp_wine2.to_excel("tmp_wine2.xlsx")

    pass

    def write_to_excel(self, experiment_list, n_features):
        # self.__write_headers(n_features)
        self.__write_colum_dataset(experiment_list, n_features)
    pass

pass